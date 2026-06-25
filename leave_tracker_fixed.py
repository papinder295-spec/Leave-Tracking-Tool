import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta
from pathlib import Path
import tempfile
import shutil
import os
import base64

# Optional Excel-open support
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except Exception:
    xw = None
    XLWINGS_AVAILABLE = False

# =============================
# PAGE CONFIG  ← MUST BE FIRST
# =============================
st.set_page_config(
    page_title="Data Estate Relationship Team Leave Tracker (Gurgaon)",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================
# CONFIG
# =============================
APP_DIR = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()

TRACKER_FILE_NAME = "Gurgaon Leaves Tracker 2026.xlsx"
PREF_FILE_NAME    = "Preferred Holidays_Gurugram.xlsx"
PREFERRED_SHEET   = "GGN_2026"
BACKGROUND_IMAGE_FILE = "moody_background.png"

DEFAULT_BASE_DIR = APP_DIR

LEAVE_ONLY_TYPES = [
    "PL", "PL 1/2",
    "SL", "SL 1/2",
    "PH", "PAT", "ED", "BL", "LWP", "SH"
]

# Present is first/default so wrong leave can be corrected directly.
LEAVE_TYPES = ["Present"] + LEAVE_ONLY_TYPES

# =============================
# BACKGROUND IMAGE HELPER
# =============================
def get_base64_image(image_path: Path) -> str | None:
    try:
        if image_path.exists():
            with open(image_path, "rb") as img_file:
                return base64.b64encode(img_file.read()).decode()
    except Exception:
        pass
    return None


bg_image_path = APP_DIR / BACKGROUND_IMAGE_FILE
bg_base64     = get_base64_image(bg_image_path)

# =============================
# STYLING
# =============================
if bg_base64:
    background_css = f"""
    background-image:
        linear-gradient(rgba(3,7,18,0.74), rgba(3,7,18,0.88)),
        url("data:image/png;base64,{bg_base64}");
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    background-attachment: fixed;
    """
else:
    background_css = (
        "background: linear-gradient(135deg, #020617 0%, #071B70 45%, #020617 100%);"
    )

st.markdown(f"""
<style>
.stApp {{
    {background_css}
}}
header[data-testid="stHeader"] {{
    height: 0rem !important;
    visibility: hidden !important;
}}
.block-container {{
    padding-top: 3.6rem !important;
    padding-bottom: 2rem !important;
    padding-left: 2.4rem !important;
    padding-right: 2.4rem !important;
    max-width: 100% !important;
    overflow: visible !important;
}}
section[data-testid="stSidebar"] {{
    background: rgba(2,6,23,0.94) !important;
    border-right: 1px solid rgba(255,255,255,0.14);
    padding-top: 1rem;
}}
.main-title {{
    font-size: clamp(28px, 2.8vw, 42px);
    font-weight: 900;
    color: #FFFFFF;
    line-height: 1.35;
    margin-top: 10px;
    margin-bottom: 8px;
    padding-top: 8px;
    padding-bottom: 4px;
    white-space: normal;
    overflow: visible !important;
    text-shadow: 0px 2px 12px rgba(0,0,0,0.70);
}}
.subtle {{
    color: #D8E6FF;
    font-size: 15px;
    line-height: 1.55;
    margin-bottom: 24px;
    overflow: visible !important;
    text-shadow: 0px 1px 8px rgba(0,0,0,0.55);
}}
[data-testid="stForm"],
[data-testid="stDataFrame"],
div[data-testid="stAlert"] {{
    background: rgba(3,7,18,0.77) !important;
    border-radius: 12px !important;
    border: 1px solid rgba(255,255,255,0.14) !important;
}}
[data-testid="stMetric"] {{
    background: rgba(3,7,18,0.74);
    border: 1px solid rgba(255,255,255,0.14);
    border-radius: 14px;
    padding: 14px 18px;
}}
h1,h2,h3,h4,h5,h6,div,span,p,label {{ overflow: visible !important; }}
h1,h2,h3,h4,h5,h6,p,label,span {{ color: #FFFFFF; }}
.stMarkdown,.stText,.stCaption {{ color: #FFFFFF !important; }}
div[data-baseweb="select"] > div, input, textarea {{
    background-color: rgba(15,23,42,0.94) !important;
    color: white !important;
    border-color: rgba(255,255,255,0.22) !important;
}}
div[data-baseweb="popover"] {{ color: white !important; z-index: 999999 !important; }}
.stButton > button, button[kind="primary"], button[kind="secondary"] {{
    background: rgba(37,99,235,0.94) !important;
    color: white !important;
    border: 1px solid rgba(255,255,255,0.25) !important;
    border-radius: 10px !important;
}}
.stButton > button:hover {{
    background: rgba(59,130,246,1) !important;
    border-color: rgba(255,255,255,0.45) !important;
}}
[data-testid="stDataFrame"] {{ margin-top: 8px; }}
div[data-testid="stAlert"] {{ margin-top: 10px; margin-bottom: 10px; }}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<div class="main-title">📊 Data Estate Relationship Team Leave Tracker (Gurgaon)</div>',
    unsafe_allow_html=True
)
st.markdown(
    '<div class="subtle">Data Estate Relationship Team Leave Tracker with correction mode, '
    'employee leave history, employee management, filters, and open-Excel support.</div>',
    unsafe_allow_html=True
)

# =============================
# HELPERS
# =============================
def get_paths(base_dir: str):
    base = Path(base_dir).expanduser()
    return base / TRACKER_FILE_NAME, base / PREF_FILE_NAME


def is_emp_id(val) -> bool:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    sval = str(val).strip()
    if sval.endswith(".0"):
        sval = sval[:-2]
    return sval.isdigit()


def normalize_emp_id(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    sval = str(val).strip()
    return sval[:-2] if sval.endswith(".0") else sval


def normalize_status(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def format_joining_date(val) -> str:
    dt = pd.to_datetime(val, errors="coerce")
    if pd.isna(dt):
        return "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val)
    return dt.strftime("%Y-%m-%d")


def is_weekend_date(d: date) -> bool:
    return d.weekday() >= 5  # 5=Sat, 6=Sun


def status_weight(status: str) -> float:
    s = normalize_status(status)
    if s in ("PL", "SL", "PH", "PAT", "ED", "BL", "LWP", "SH"):
        return 1.0
    if s in ("PL 1/2", "SL 1/2"):
        return 0.5
    return 0.0


def daterange(start_date: date, end_date: date):
    d = start_date
    while d <= end_date:
        yield d
        d += timedelta(days=1)


def safe_save_workbook(wb, file_path: Path):
    file_path = Path(file_path)
    tmp_fd, tmp_name = tempfile.mkstemp(
        suffix=file_path.suffix,
        dir=str(file_path.parent)
    )
    os.close(tmp_fd)
    try:
        wb.save(tmp_name)
        shutil.move(tmp_name, str(file_path))
    finally:
        if os.path.exists(tmp_name):
            try:
                os.remove(tmp_name)
            except Exception:
                pass


def get_open_xlwings_book(file_path: Path):
    """Return the xlwings Book object if the file is already open in Excel, else None."""
    if not XLWINGS_AVAILABLE:
        return None
    target_name = Path(file_path).name.lower()
    try:
        for app in xw.apps:
            try:
                for book in app.books:
                    try:
                        if Path(book.fullname).name.lower() == target_name:
                            return book
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception:
        pass
    return None


def get_tracker_header_map_openpyxl(ws):
    """Return (emp_col_index, {date: col_index}) from row-1 headers."""
    header_map: dict[date, int] = {}
    emp_col: int | None = None
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None and str(value).strip() == "Emp ID":
            emp_col = col
        dt = pd.to_datetime(value, errors="coerce")
        if pd.notna(dt):
            header_map[dt.date()] = col
    return emp_col, header_map


def get_tracker_emp_row_openpyxl(ws, emp_id: str):
    """Return (row_index, emp_col_index) for the given emp_id."""
    emp_col, _ = get_tracker_header_map_openpyxl(ws)
    if emp_col is None:
        raise ValueError("'Emp ID' column was not found in the tracker workbook.")
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=emp_col).value
        if is_emp_id(val) and normalize_emp_id(val) == str(emp_id):
            return row, emp_col
    raise ValueError(f"Emp ID {emp_id} was not found in the tracker workbook.")


def get_pref_header_info_openpyxl(ws):
    """Return (header_row, col_map, date_map) for the preferred-holidays sheet."""
    header_row = 2
    col_map: dict[str, int] = {}
    date_map: dict[date, int] = {}
    KNOWN_COLS = {
        "S. No", "Emp_Code", "Employee Name", "Total",
        "Joining Date", "Function", "Reporting Manager", "Team"
    }
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        sval = str(val).strip() if val is not None else ""
        if sval in KNOWN_COLS:
            col_map[sval] = col
        dt = pd.to_datetime(val, errors="coerce")
        if pd.notna(dt):
            date_map[dt.date()] = col
    return header_row, col_map, date_map


# =============================
# DATA LOADERS
# =============================
@st.cache_data(show_spinner=False)
def load_tracker_data(tracker_path_str: str):
    df = pd.read_excel(tracker_path_str, engine="openpyxl")

    if "Emp ID" not in df.columns:
        raise ValueError("'Emp ID' column was not found in the tracker file.")
    if "Name" not in df.columns:
        raise ValueError("'Name' column was not found in the tracker file.")

    df = df[df["Emp ID"].apply(is_emp_id)].copy()
    df["Emp ID"] = df["Emp ID"].apply(normalize_emp_id)
    df["Name"]   = df["Name"].astype(str).str.strip()

    if "Joining Date" in df.columns:
        df["Joining Date"] = df["Joining Date"].apply(format_joining_date)

    # Ensure optional columns exist
    for col in ("Team", "Designation"):
        if col not in df.columns:
            df[col] = ""

    STANDARD_COLS = {"Emp ID", "Name", "Joining Date", "Team", "Designation"}
    id_vars   = [c for c in df.columns if c in STANDARD_COLS]
    date_cols = {}
    for c in df.columns:
        if c in STANDARD_COLS:
            continue
        dt = pd.to_datetime(c, errors="coerce")
        if pd.notna(dt):
            date_cols[c] = dt.date()

    long_df = df.melt(
        id_vars=id_vars,
        value_vars=list(date_cols.keys()),
        var_name="DateCol",
        value_name="Status"
    )
    long_df["Date"]         = long_df["DateCol"].map(date_cols)
    long_df["Status"]       = long_df["Status"].apply(normalize_status)
    long_df["Leave Weight"] = long_df["Status"].apply(status_weight)
    long_df.drop(columns=["DateCol"], inplace=True)

    return df, long_df


@st.cache_data(show_spinner=False)
def load_preferred_data(pref_path_str: str):
    pref_df = pd.read_excel(
        pref_path_str,
        sheet_name=PREFERRED_SHEET,
        engine="openpyxl",
        header=1
    )
    if "Emp_Code" not in pref_df.columns:
        return pd.DataFrame()

    pref_df = pref_df[pref_df["Emp_Code"].apply(is_emp_id)].copy()
    pref_df["Emp_Code"] = pref_df["Emp_Code"].apply(normalize_emp_id)

    if "Employee Name" in pref_df.columns:
        pref_df["Employee Name"] = pref_df["Employee Name"].astype(str).str.strip()
    if "Joining Date" in pref_df.columns:
        pref_df["Joining Date"] = pref_df["Joining Date"].apply(format_joining_date)

    return pref_df


def clear_caches():
    load_tracker_data.clear()
    load_preferred_data.clear()


# =============================
# BUSINESS LOGIC
# =============================
def get_emp_summary(long_df: pd.DataFrame) -> pd.DataFrame:
    def cnt_exact(series, target):
        return sum(1 for v in series if normalize_status(v) == target)

    for col in ("Team", "Designation"):
        if col not in long_df.columns:
            long_df[col] = ""

    summary = long_df.groupby(
        ["Emp ID", "Name", "Team", "Designation"],
        dropna=False
    ).agg(
        PL          =("Status", lambda x: cnt_exact(x, "PL")),
        PL_Half     =("Status", lambda x: cnt_exact(x, "PL 1/2")),
        SL          =("Status", lambda x: cnt_exact(x, "SL")),
        SL_Half     =("Status", lambda x: cnt_exact(x, "SL 1/2")),
        PH          =("Status", lambda x: cnt_exact(x, "PH")),
        PAT         =("Status", lambda x: cnt_exact(x, "PAT")),
        ED          =("Status", lambda x: cnt_exact(x, "ED")),
        BL          =("Status", lambda x: cnt_exact(x, "BL")),
        LWP         =("Status", lambda x: cnt_exact(x, "LWP")),
        SH          =("Status", lambda x: cnt_exact(x, "SH")),
        Total_Leave_Days=("Leave Weight", "sum"),
    ).reset_index()

    return summary.sort_values("Name")


def update_tracker_workbook(
    tracker_path: Path,
    emp_id: str,
    start_date: date,
    end_date: date,
    leave_type: str,
):
    missed_dates:   list[str] = []
    skipped_weekends: list[str] = []
    updated_dates:  list[str] = []
    changes_log:    list[dict] = []

    open_book = get_open_xlwings_book(tracker_path)

    # --- Mode 1: workbook already open in Excel (xlwings) ---
    if open_book is not None:
        ws     = open_book.sheets[0]
        values = ws.used_range.value

        if not values or len(values) < 2:
            raise ValueError("The tracker workbook does not contain valid data.")

        headers    = values[0]
        header_map: dict[date, int] = {}
        emp_col: int | None = None

        for idx, h in enumerate(headers, start=1):
            if h is not None and str(h).strip() == "Emp ID":
                emp_col = idx
            dt = pd.to_datetime(h, errors="coerce")
            if pd.notna(dt):
                header_map[dt.date()] = idx

        if emp_col is None:
            raise ValueError("'Emp ID' column was not found in the tracker workbook.")

        emp_row: int | None = None
        for r in range(2, len(values) + 1):
            row_vals = values[r - 1]
            cell_val = row_vals[emp_col - 1] if row_vals else None
            if is_emp_id(cell_val) and normalize_emp_id(cell_val) == str(emp_id):
                emp_row = r
                break

        if emp_row is None:
            raise ValueError(f"Emp ID {emp_id} was not found in the tracker workbook.")

        for d in daterange(start_date, end_date):
            if is_weekend_date(d):
                skipped_weekends.append(str(d))
                continue
            col = header_map.get(d)
            if col is None:
                missed_dates.append(str(d))
                continue

            old_val = normalize_status(ws.range((emp_row, col)).value)
            new_val = "" if leave_type == "Present" else leave_type
            ws.range((emp_row, col)).value = new_val

            updated_dates.append(str(d))
            changes_log.append({
                "emp_id":    str(emp_id),
                "date":      d,
                "old_value": old_val,
                "new_value": normalize_status(new_val),
            })

        open_book.save()
        return missed_dates, skipped_weekends, updated_dates, changes_log

    # --- Mode 2: workbook closed → openpyxl fallback ---
    wb = load_workbook(tracker_path)
    ws = wb[wb.sheetnames[0]]

    emp_row, _  = get_tracker_emp_row_openpyxl(ws, emp_id)
    _, header_map = get_tracker_header_map_openpyxl(ws)

    for d in daterange(start_date, end_date):
        if is_weekend_date(d):
            skipped_weekends.append(str(d))
            continue
        col = header_map.get(d)
        if col is None:
            missed_dates.append(str(d))
            continue

        old_val = normalize_status(ws.cell(row=emp_row, column=col).value)
        new_val = "" if leave_type == "Present" else leave_type
        ws.cell(row=emp_row, column=col).value = new_val

        updated_dates.append(str(d))
        changes_log.append({
            "emp_id":    str(emp_id),
            "date":      d,
            "old_value": old_val,
            "new_value": normalize_status(new_val),
        })

    safe_save_workbook(wb, tracker_path)
    return missed_dates, skipped_weekends, updated_dates, changes_log


def apply_preferred_marker(
    pref_path: Path,
    emp_id: str,
    target_date: date,
    mark_y: bool,
) -> bool:
    if not pref_path.exists():
        return False

    open_book = get_open_xlwings_book(pref_path)

    # --- xlwings mode ---
    if open_book is not None:
        try:
            ws = open_book.sheets[PREFERRED_SHEET]
        except Exception:
            return False

        values = ws.used_range.value
        if not values or len(values) < 3:
            return False

        header_row = 2
        headers    = values[header_row - 1]

        emp_code_col: int | None = None
        total_col:    int | None = None
        date_col:     int | None = None

        for idx, h in enumerate(headers, start=1):
            if h == "Emp_Code":
                emp_code_col = idx
            elif h == "Total":
                total_col = idx
            else:
                dt = pd.to_datetime(h, errors="coerce")
                if pd.notna(dt) and dt.date() == target_date:
                    date_col = idx

        if emp_code_col is None or date_col is None:
            return False

        emp_row: int | None = None
        for r in range(header_row + 1, len(values) + 1):
            row_vals = values[r - 1]
            cell_val = row_vals[emp_code_col - 1] if row_vals else None
            if is_emp_id(cell_val) and normalize_emp_id(cell_val) == str(emp_id):
                emp_row = r
                break

        if emp_row is None:
            return False

        ws.range((emp_row, date_col)).value = "Y" if mark_y else None

        if total_col is not None:
            y_count = sum(
                1
                for c, h in enumerate(headers, start=1)
                if pd.notna(pd.to_datetime(h, errors="coerce"))
                and str(ws.range((emp_row, c)).value).strip().upper() == "Y"
            )
            ws.range((emp_row, total_col)).value = y_count

        open_book.save()
        return True

    # --- openpyxl mode ---
    wb = load_workbook(pref_path)
    if PREFERRED_SHEET not in wb.sheetnames:
        return False

    ws = wb[PREFERRED_SHEET]
    header_row, col_map, date_map = get_pref_header_info_openpyxl(ws)

    emp_col  = col_map.get("Emp_Code")
    date_col = date_map.get(target_date)
    if emp_col is None or date_col is None:
        return False

    emp_row: int | None = None
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=emp_col).value
        if is_emp_id(val) and normalize_emp_id(val) == str(emp_id):
            emp_row = row
            break

    if emp_row is None:
        return False

    ws.cell(row=emp_row, column=date_col).value = "Y" if mark_y else None

    if "Total" in col_map:
        total_col = col_map["Total"]
        y_count = sum(
            1
            for d_col in date_map.values()
            if str(ws.cell(row=emp_row, column=d_col).value).strip().upper() == "Y"
        )
        ws.cell(row=emp_row, column=total_col).value = y_count

    safe_save_workbook(wb, pref_path)
    return True


def sync_preferred_from_change_log(
    pref_path: Path,
    change_log: list[dict],
    use_new_value: bool = True,
):
    if not pref_path.exists():
        return
    for ch in change_log:
        # Weekends already excluded at update time; guard anyway
        if is_weekend_date(ch["date"]):
            continue
        status = ch["new_value"] if use_new_value else ch["old_value"]
        mark_y = normalize_status(status) == "PH"
        apply_preferred_marker(pref_path, ch["emp_id"], ch["date"], mark_y)


def add_employee_to_tracker(
    tracker_path: Path,
    emp_id: str,
    name: str,
    joining_date_value: str,
    team: str,
    designation: str,
):
    open_book = get_open_xlwings_book(tracker_path)
    REQUIRED = ["Emp ID", "Name", "Joining Date", "Team", "Designation"]

    if open_book is not None:
        ws     = open_book.sheets[0]
        values = ws.used_range.value
        headers = values[0]
        lookup = {str(h).strip(): idx for idx, h in enumerate(headers, start=1)}

        for col in REQUIRED:
            if col not in lookup:
                raise ValueError(f"'{col}' column was not found in the tracker workbook.")

        for r in range(2, len(values) + 1):
            row_vals = values[r - 1]
            existing = row_vals[lookup["Emp ID"] - 1] if row_vals else None
            if is_emp_id(existing) and normalize_emp_id(existing) == str(emp_id):
                raise ValueError(f"Emp ID {emp_id} already exists in the tracker workbook.")

        # FIX: xlwings used_range.value is 0-indexed length; next empty row = len(values) + 1
        next_row = len(values) + 1
        ws.range((next_row, lookup["Emp ID"])).value         = str(emp_id)
        ws.range((next_row, lookup["Name"])).value           = name
        ws.range((next_row, lookup["Joining Date"])).value   = joining_date_value
        ws.range((next_row, lookup["Team"])).value           = team
        ws.range((next_row, lookup["Designation"])).value    = designation
        open_book.save()
        return

    wb = load_workbook(tracker_path)
    ws = wb[wb.sheetnames[0]]
    header_lookup = {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
    }

    for col in REQUIRED:
        if col not in header_lookup:
            raise ValueError(f"'{col}' column was not found in the tracker workbook.")

    emp_col = header_lookup["Emp ID"]
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=emp_col).value
        if is_emp_id(val) and normalize_emp_id(val) == str(emp_id):
            raise ValueError(f"Emp ID {emp_id} already exists in the tracker workbook.")

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=header_lookup["Emp ID"]).value         = str(emp_id)
    ws.cell(row=next_row, column=header_lookup["Name"]).value           = name
    ws.cell(row=next_row, column=header_lookup["Joining Date"]).value   = joining_date_value
    ws.cell(row=next_row, column=header_lookup["Team"]).value           = team
    ws.cell(row=next_row, column=header_lookup["Designation"]).value    = designation
    safe_save_workbook(wb, tracker_path)


def add_employee_to_preferred(
    pref_path: Path,
    emp_id: str,
    name: str,
    joining_date_value: str,
    team: str,
):
    if not pref_path.exists():
        return

    open_book = get_open_xlwings_book(pref_path)

    if open_book is not None:
        try:
            ws = open_book.sheets[PREFERRED_SHEET]
        except Exception:
            return

        values      = ws.used_range.value
        header_row  = 2
        headers     = values[header_row - 1]
        lookup      = {str(h).strip(): idx for idx, h in enumerate(headers, start=1)}

        emp_col = lookup.get("Emp_Code")
        if emp_col is None:
            return

        for r in range(header_row + 1, len(values) + 1):
            row_vals = values[r - 1]
            existing = row_vals[emp_col - 1] if row_vals else None
            if is_emp_id(existing) and normalize_emp_id(existing) == str(emp_id):
                return  # already exists, skip silently

        # FIX: correct next-row calculation for xlwings
        next_row = len(values) + 1
        _set = lambda key, val: ws.range((next_row, lookup[key])).value.__setattr__ or \
               ws.range((next_row, lookup[key])).__setattr__  # noqa – use direct assignment below

        if "S. No"            in lookup: ws.range((next_row, lookup["S. No"])).value           = max(1, next_row - header_row)
        if "Emp_Code"         in lookup: ws.range((next_row, lookup["Emp_Code"])).value         = str(emp_id)
        if "Employee Name"    in lookup: ws.range((next_row, lookup["Employee Name"])).value    = name
        if "Total"            in lookup: ws.range((next_row, lookup["Total"])).value            = 0
        if "Joining Date"     in lookup: ws.range((next_row, lookup["Joining Date"])).value     = joining_date_value
        if "Function"         in lookup: ws.range((next_row, lookup["Function"])).value         = ""
        if "Reporting Manager" in lookup: ws.range((next_row, lookup["Reporting Manager"])).value = ""
        if "Team"             in lookup: ws.range((next_row, lookup["Team"])).value             = team
        open_book.save()
        return

    wb = load_workbook(pref_path)
    if PREFERRED_SHEET not in wb.sheetnames:
        return

    ws = wb[PREFERRED_SHEET]
    header_row, col_map, _ = get_pref_header_info_openpyxl(ws)

    emp_col = col_map.get("Emp_Code")
    if emp_col is None:
        return

    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=emp_col).value
        if is_emp_id(val) and normalize_emp_id(val) == str(emp_id):
            return  # already exists

    next_row = ws.max_row + 1
    if "S. No"             in col_map: ws.cell(row=next_row, column=col_map["S. No"]).value            = max(1, next_row - header_row)
    if "Emp_Code"          in col_map: ws.cell(row=next_row, column=col_map["Emp_Code"]).value          = str(emp_id)
    if "Employee Name"     in col_map: ws.cell(row=next_row, column=col_map["Employee Name"]).value     = name
    if "Total"             in col_map: ws.cell(row=next_row, column=col_map["Total"]).value             = 0
    if "Joining Date"      in col_map: ws.cell(row=next_row, column=col_map["Joining Date"]).value      = joining_date_value
    if "Function"          in col_map: ws.cell(row=next_row, column=col_map["Function"]).value          = ""
    if "Reporting Manager" in col_map: ws.cell(row=next_row, column=col_map["Reporting Manager"]).value = ""
    if "Team"              in col_map: ws.cell(row=next_row, column=col_map["Team"]).value              = team
    safe_save_workbook(wb, pref_path)


def remove_employee_from_tracker(tracker_path: Path, emp_id: str):
    open_book = get_open_xlwings_book(tracker_path)

    if open_book is not None:
        ws      = open_book.sheets[0]
        values  = ws.used_range.value
        headers = values[0]

        emp_col: int | None = None
        for idx, h in enumerate(headers, start=1):
            if h is not None and str(h).strip() == "Emp ID":
                emp_col = idx
                break

        if emp_col is None:
            raise ValueError("'Emp ID' column was not found in the tracker workbook.")

        delete_row: int | None = None
        for r in range(2, len(values) + 1):
            row_vals = values[r - 1]
            cell_val = row_vals[emp_col - 1] if row_vals else None
            if is_emp_id(cell_val) and normalize_emp_id(cell_val) == str(emp_id):
                delete_row = r
                break

        if delete_row is None:
            raise ValueError(f"Emp ID {emp_id} was not found in the tracker workbook.")

        ws.api.Rows(delete_row).Delete()
        open_book.save()
        return

    wb = load_workbook(tracker_path)
    ws = wb[wb.sheetnames[0]]
    emp_row, _ = get_tracker_emp_row_openpyxl(ws, emp_id)
    ws.delete_rows(emp_row, 1)
    safe_save_workbook(wb, tracker_path)


def remove_employee_from_preferred(pref_path: Path, emp_id: str):
    if not pref_path.exists():
        return

    open_book = get_open_xlwings_book(pref_path)

    if open_book is not None:
        try:
            ws = open_book.sheets[PREFERRED_SHEET]
        except Exception:
            return

        values     = ws.used_range.value
        header_row = 2
        headers    = values[header_row - 1]

        emp_col: int | None = None
        for idx, h in enumerate(headers, start=1):
            if h is not None and str(h).strip() == "Emp_Code":
                emp_col = idx
                break

        if emp_col is None:
            return

        delete_row: int | None = None
        for r in range(header_row + 1, len(values) + 1):
            row_vals = values[r - 1]
            cell_val = row_vals[emp_col - 1] if row_vals else None
            if is_emp_id(cell_val) and normalize_emp_id(cell_val) == str(emp_id):
                delete_row = r
                break

        if delete_row is None:
            return  # not found → silent (consistent with add behaviour)

        ws.api.Rows(delete_row).Delete()
        open_book.save()
        return

    wb = load_workbook(pref_path)
    if PREFERRED_SHEET not in wb.sheetnames:
        return

    ws = wb[PREFERRED_SHEET]
    header_row, col_map, _ = get_pref_header_info_openpyxl(ws)
    emp_col = col_map.get("Emp_Code")
    if emp_col is None:
        return

    delete_row: int | None = None
    for row in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=emp_col).value
        if is_emp_id(val) and normalize_emp_id(val) == str(emp_id):
            delete_row = row
            break

    if delete_row is None:
        return

    ws.delete_rows(delete_row, 1)
    safe_save_workbook(wb, pref_path)


# =============================
# SIDEBAR
# =============================
st.sidebar.header("🛠️ Admin Control Center")

base_dir = st.sidebar.text_input("Data folder path", str(DEFAULT_BASE_DIR))
tracker_path, pref_path = get_paths(base_dir)

menu = st.sidebar.radio(
    "Navigate Menu",
    [
        "Live Dashboard",
        "Log / Update Leave",
        "Employee Leave History",
        "Employee Management",
        "Employee Directory",
        "Leave Summary",
    ],
)

selected_date          = st.sidebar.date_input("📅 Date Check", value=date.today())
team_filter_sidebar    = st.sidebar.text_input("Team filter (optional)", "")
employee_search_sidebar = st.sidebar.text_input("Employee search (optional)", "")

# =============================
# FILE VALIDATION + DATA LOAD
# =============================
if not tracker_path.exists():
    st.error(f"Tracker file was not found:\n{tracker_path}")
    st.stop()

try:
    tracker_df, long_df = load_tracker_data(str(tracker_path))
except Exception as e:
    st.error(f"Tracker file could not be loaded: {e}")
    st.stop()

try:
    pref_df = load_preferred_data(str(pref_path)) if pref_path.exists() else pd.DataFrame()
except Exception:
    pref_df = pd.DataFrame()

summary_df = get_emp_summary(long_df)

# Ensure optional columns exist in all frames
for col in ("Team", "Designation", "Joining Date"):
    if col not in tracker_df.columns:
        tracker_df[col] = ""

for col in ("Team", "Designation"):
    if col not in long_df.columns:
        long_df[col] = ""
    if col not in summary_df.columns:
        summary_df[col] = ""

# Optional sidebar filters
if team_filter_sidebar.strip():
    _tf = team_filter_sidebar
    tracker_df = tracker_df[tracker_df["Team"].astype(str).str.contains(_tf, case=False, na=False)].copy()
    long_df    = long_df[long_df["Team"].astype(str).str.contains(_tf, case=False, na=False)].copy()
    summary_df = summary_df[summary_df["Team"].astype(str).str.contains(_tf, case=False, na=False)].copy()

if employee_search_sidebar.strip():
    _es = employee_search_sidebar
    tracker_df = tracker_df[tracker_df["Name"].astype(str).str.contains(_es, case=False, na=False)].copy()
    long_df    = long_df[long_df["Name"].astype(str).str.contains(_es, case=False, na=False)].copy()
    summary_df = summary_df[summary_df["Name"].astype(str).str.contains(_es, case=False, na=False)].copy()

# =============================
# LIVE DASHBOARD
# =============================
if menu == "Live Dashboard":
    day_df = long_df[long_df["Date"] == selected_date].copy()

    if day_df.empty:
        merged = tracker_df[["Emp ID", "Name", "Team", "Designation"]].copy()
        merged["Status"] = ""
    else:
        merged = day_df[["Emp ID", "Name", "Team", "Designation", "Status"]].copy()

    is_weekend     = is_weekend_date(selected_date)
    weekend_label  = "WO" if is_weekend else "Present"
    merged["Status Today"] = merged["Status"].replace("", weekend_label)

    on_leave_df      = merged[merged["Status Today"].isin(LEAVE_ONLY_TYPES)].copy()
    total_staff      = len(merged)
    on_leave_count   = int(on_leave_df.shape[0])
    weekend_off_count = int((merged["Status Today"] == "WO").sum())
    present_count    = int((merged["Status Today"] == "Present").sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Staff",       total_staff)
    c2.metric("Present Today",     present_count)
    c3.metric("On Leave Today",    on_leave_count)
    c4.metric("Weekend Off Today", weekend_off_count)

    st.markdown("### 👥 People On Leave Today")
    if on_leave_df.empty:
        st.info("No employees are on leave for the selected date.")
    else:
        leave_display = on_leave_df[["Name", "Status Today", "Team", "Designation"]].copy()
        leave_display.columns = ["Employee Name", "Leave Type", "Team", "Designation"]
        st.dataframe(
            leave_display.sort_values(["Leave Type", "Employee Name"]),
            use_container_width=True,
            hide_index=True,
        )

# =============================
# LOG / UPDATE LEAVE
# =============================
elif menu == "Log / Update Leave":
    st.markdown("### 📝 Log / Update Leave")
    st.caption(
        "Use **Present** to remove/correct an incorrect leave entry. "
        "Select employee → keep Leave Type = Present → choose date → Save."
    )

    if tracker_df.empty:
        st.info("No employee data is available.")
    else:
        # FIX: selectbox/date_input must be OUTSIDE the form when their
        # value is read before form submission; here they are correctly
        # inside the form and only read on submit — this is fine.
        with st.form("leave_form", clear_on_submit=False):
            emp_name = st.selectbox(
                "Select Employee",
                tracker_df["Name"].dropna().sort_values().tolist(),
            )
            leave_type = st.selectbox(
                "Leave Type",
                LEAVE_TYPES,
                index=0,
            )
            d_start    = st.date_input("Start Date", value=date.today())
            d_end      = st.date_input("End Date",   value=date.today())
            submitted  = st.form_submit_button("💾 Save")

        st.markdown("#### Correction Guidelines")
        st.markdown(
            "- To remove wrong leave → select employee, same date range, Leave Type = Present, Save.\n"
            "- **Present** clears the cell and treats the employee as present.\n"
            "- Weekends are automatically skipped.\n"
            "- Dashboard and Leave History refresh automatically after save."
        )

        if submitted:
            if d_start > d_end:
                st.error("Start Date cannot be later than End Date.")
            else:
                emp_row_data = tracker_df[tracker_df["Name"] == emp_name]
                if emp_row_data.empty:
                    st.error(f"Employee '{emp_name}' not found.")
                else:
                    emp_id = emp_row_data.iloc[0]["Emp ID"]
                    try:
                        missed, skipped, updated, change_log = update_tracker_workbook(
                            tracker_path=tracker_path,
                            emp_id=str(emp_id),
                            start_date=d_start,
                            end_date=d_end,
                            leave_type=leave_type,
                        )

                        if pref_path.exists() and change_log:
                            sync_preferred_from_change_log(pref_path, change_log, use_new_value=True)

                        clear_caches()
                        st.success(f"✅ Leave updated for **{emp_name}**.")

                        if updated:
                            st.write(
                                "Updated dates:",
                                ", ".join(updated[:15]) + (" …" if len(updated) > 15 else ""),
                            )
                        if skipped:
                            st.warning(
                                "Weekend dates skipped: "
                                + ", ".join(skipped[:10])
                                + (" …" if len(skipped) > 10 else "")
                            )
                        if missed:
                            st.warning(
                                "Dates not found in tracker columns: "
                                + ", ".join(missed)
                            )

                        st.rerun()

                    except Exception as e:
                        st.error(f"Update failed: {e}")

# =============================
# EMPLOYEE LEAVE HISTORY
# =============================
elif menu == "Employee Leave History":
    st.markdown("### 🕘 Employee Leave History")

    if tracker_df.empty:
        st.info("No employee data is available.")
    else:
        chosen_name = st.selectbox(
            "Select Employee",
            tracker_df["Name"].dropna().sort_values().tolist(),
            key="hist_emp",
        )

        hist_df = long_df[
            (long_df["Name"] == chosen_name) &
            (long_df["Status"].isin(LEAVE_ONLY_TYPES))
        ].copy()

        if hist_df.empty:
            st.info("No leave history available for the selected employee.")
        else:
            hist_df = hist_df[["Date", "Status", "Team", "Designation"]].sort_values(
                "Date", ascending=False
            )
            hist_df.columns = ["Date", "Leave Type", "Team", "Designation"]
            st.dataframe(hist_df, use_container_width=True, hide_index=True)

# =============================
# EMPLOYEE MANAGEMENT
# =============================
elif menu == "Employee Management":
    st.markdown("### 👥 Employee Management")

    tab_add, tab_remove = st.tabs(["Add New Employee", "Remove Employee"])

    with tab_add:
        st.markdown("#### Add New Employee")
        st.caption(
            "The employee will also be added to the preferred holiday workbook if available."
        )

        with st.form("add_employee_form", clear_on_submit=False):
            new_emp_id      = st.text_input("Emp ID")
            new_name        = st.text_input("Name")
            new_joining_date = st.date_input("Joining Date", value=date.today())
            new_team        = st.text_input("Team")
            new_designation = st.text_input("Designation")
            add_submitted   = st.form_submit_button("➕ Add Employee")

        if add_submitted:
            if not str(new_emp_id).strip().isdigit():
                st.error("Emp ID must be numeric.")
            elif not str(new_name).strip():
                st.error("Name is required.")
            else:
                try:
                    joining_str = pd.to_datetime(new_joining_date).strftime("%Y-%m-%d")
                    add_employee_to_tracker(
                        tracker_path=tracker_path,
                        emp_id=str(new_emp_id).strip(),
                        name=str(new_name).strip(),
                        joining_date_value=joining_str,
                        team=str(new_team).strip(),
                        designation=str(new_designation).strip(),
                    )
                    if pref_path.exists():
                        add_employee_to_preferred(
                            pref_path=pref_path,
                            emp_id=str(new_emp_id).strip(),
                            name=str(new_name).strip(),
                            joining_date_value=joining_str,
                            team=str(new_team).strip(),
                        )
                    clear_caches()
                    st.success(f"✅ Employee **{new_name}** ({new_emp_id}) added successfully.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Add employee failed: {e}")

    with tab_remove:
        st.markdown("#### Remove Employee")
        st.caption("The employee will also be removed from the preferred holiday workbook if available.")

        if tracker_df.empty:
            st.info("No employees available to remove.")
        else:
            remove_name = st.selectbox(
                "Select Employee to Remove",
                tracker_df["Name"].dropna().sort_values().tolist(),
                key="remove_emp",
            )
            remove_row_data  = tracker_df[tracker_df["Name"] == remove_name]
            remove_emp_id    = remove_row_data.iloc[0]["Emp ID"] if not remove_row_data.empty else None

            st.warning(f"You are about to remove: **{remove_name}** ({remove_emp_id})")

            confirm_remove = st.checkbox(
                "I understand this will permanently remove the employee record from the tool files."
            )

            if st.button("🗑️ Remove Employee", disabled=not confirm_remove):
                if remove_emp_id is None:
                    st.error("Could not determine Emp ID.")
                else:
                    try:
                        remove_employee_from_tracker(tracker_path, str(remove_emp_id))
                        if pref_path.exists():
                            remove_employee_from_preferred(pref_path, str(remove_emp_id))
                        clear_caches()
                        st.success(f"✅ Employee **{remove_name}** ({remove_emp_id}) removed.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Remove employee failed: {e}")

# =============================
# EMPLOYEE DIRECTORY
# =============================
elif menu == "Employee Directory":
    st.markdown("### 👤 Employee Directory")

    DIRECTORY_COLS = ["Emp ID", "Name", "Joining Date", "Team", "Designation"]
    for col in DIRECTORY_COLS:
        if col not in tracker_df.columns:
            tracker_df[col] = ""

    out = tracker_df[DIRECTORY_COLS].copy()

    if (
        not pref_df.empty
        and "Reporting Manager" in pref_df.columns
        and "Emp_Code" in pref_df.columns
    ):
        pref_small = pref_df[["Emp_Code", "Reporting Manager"]].rename(
            columns={"Emp_Code": "Emp ID"}
        )
        out = out.merge(pref_small, on="Emp ID", how="left")

    st.dataframe(out.sort_values("Name"), use_container_width=True, hide_index=True)

# =============================
# LEAVE SUMMARY
# =============================
elif menu == "Leave Summary":
    st.markdown("### 📈 Leave Summary")

    SUMMARY_COLS_IN = [
        "Name", "Team", "Designation",
        "PL", "PL_Half", "SL", "SL_Half",
        "PH", "PAT", "ED", "BL", "LWP", "SH",
        "Total_Leave_Days",
    ]
    SUMMARY_COLS_OUT = [
        "Employee Name", "Team", "Designation",
        "PL", "PL 1/2", "SL", "SL 1/2",
        "PH", "PAT", "ED", "BL", "LWP", "SH",
        "Total Leave Days",
    ]

    for col in SUMMARY_COLS_IN:
        if col not in summary_df.columns:
            summary_df[col] = 0 if col not in ("Name", "Team", "Designation") else ""

    display_summary = summary_df[SUMMARY_COLS_IN].copy()
    display_summary.columns = SUMMARY_COLS_OUT

    st.dataframe(display_summary, use_container_width=True, hide_index=True)
